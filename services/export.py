import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ExportService:
    """Профессиональный экспорт в Excel с форматированием"""
    
    def __init__(self):
        self.headers = [
            'ID', 'Дата', 'Источник', 'Компания', 'ИНН', 'Телефон', 
            'Email', 'Telegram', 'Город', 'Регион', 'Тип груза', 
            'Объём', 'Описание', 'Тип лида', 'AI Score', 'Анализ', 'Ссылка'
        ]
    
    def export_to_excel(self, leads: list, user_name: str = "user") -> BytesIO:
        """Создаёт профессиональный Excel файл"""
        output = BytesIO()
        
        # Подготовка данных
        data = []
        for lead in leads:
            data.append({
                'ID': lead.get('id', ''),
                'Дата': lead.get('created_at', '').strftime('%Y-%m-%d %H:%M') if isinstance(lead.get('created_at'), datetime) else lead.get('created_at', ''),
                'Источник': lead.get('source', ''),
                'Компания': lead.get('company', ''),
                'ИНН': lead.get('inn', ''),
                'Телефон': lead.get('phone', ''),
                'Email': lead.get('email', ''),
                'Telegram': lead.get('telegram', ''),
                'Город': lead.get('city', ''),
                'Регион': lead.get('region', ''),
                'Тип груза': lead.get('cargo_type', ''),
                'Объём': lead.get('volume', ''),
                'Описание': lead.get('description', '')[:255],
                'Тип лида': '🔥 Горячий' if lead.get('lead_type') == 'hot' else '🟡 Тёплый' if lead.get('lead_type') == 'warm' else '⚪ Холодный',
                'AI Score': lead.get('ai_score', ''),
                'Анализ': lead.get('ai_analysis', '')[:100],
                'Ссылка': lead.get('source_url', ''),
            })
        
        df = pd.DataFrame(data, columns=self.headers)
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Лиды')
            worksheet = writer.sheets['Лиды']
            
            # Форматирование
            self._apply_formatting(worksheet, len(df))
            
            # Добавляем лист со статистикой
            self._add_stats_sheet(writer, leads)
        
        output.seek(0)
        logger.info(f"Excel exported: {len(leads)} leads")
        return output
    
    def _apply_formatting(self, worksheet, row_count: int):
        """Применяет профессиональное форматирование"""
        # Заголовки
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(wrap_text=True, vertical="center")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Цветовая кодировка типов лидов
        hot_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Красный
        warm_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Жёлтый
        
        for row in range(2, row_count + 2):
            lead_type = worksheet[f'O{row}'].value  # Колонка "Тип лида"
            if lead_type and 'Горячий' in lead_type:
                worksheet[f'O{row}'].fill = hot_fill
            elif lead_type and 'Тёплый' in lead_type:
                worksheet[f'O{row}'].fill = warm_fill
    
    def _add_stats_sheet(self, writer, leads: list):
        """Добавляет лист со сводной статистикой"""
        stats_ws = writer.book.create_sheet('Статистика')
        
        total = len(leads)
        hot = sum(1 for l in leads if l.get('lead_type') == 'hot')
        warm = sum(1 for l in leads if l.get('lead_type') == 'warm')
        with_contacts = sum(1 for l in leads if l.get('phone') or l.get('telegram'))
        
        stats = [
            ['МЕТРИКА', 'ЗНАЧЕНИЕ'],
            ['Всего лидов', total],
            ['🔥 Горячие', hot],
            ['🟡 Тёплые', warm],
            ['⚪ Холодные', total - hot - warm],
            ['С контактами', with_contacts],
            ['Конверсия в контакт', f'{with_contacts/total*100:.1f}%' if total > 0 else '0%'],
            ['Дата экспорта', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ]
        
        for row_idx, row_data in enumerate(stats, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = stats_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True)
        
        stats_ws.column_dimensions['A'].width = 25
        stats_ws.column_dimensions['B'].width = 15
